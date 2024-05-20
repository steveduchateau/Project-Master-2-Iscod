from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import openpyxl
import os
from concurrent.futures import ThreadPoolExecutor

# Ouvrir le fichier Excel contenant les lettres de commande
excel_file = '/Users/steveduchateaumelaniesteve/Desktop/liste LC LHH .xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb['LC']

# Chemin du fichier de sauvegarde
backup_file = 'last_successful_iteration.txt'

# Réinitialiser last_successful_iteration à 0
last_successful_iteration = 0

# Créer un nouveau classeur Excel et ajouter une feuille (une seule fois avant la boucle while)
workbook = openpyxl.Workbook()
sheet = workbook.active

# Écrire les en-têtes dans la feuille Excel (une seule fois avant la boucle while)
sheet['A1'] = 'Lettre de Commande'
sheet['B1'] = 'Prénom et Nom'

# Définir la ligne de départ
start_iteration = 2

# Traiter toutes les lignes du fichier Excel à partir de la ligne spécifiée
max_iterations = ws.max_row

def process_iteration(current_iteration):
    print(f"*** Début de l'itération {current_iteration} ***")

    try:
        # Initialise le navigateur avec une attente implicite de 30 secondes
        driver = webdriver.Chrome()
        driver.implicitly_wait(30)

        # Ouvrir l'URL
        driver.get("https://www.portail-emploi.fr")

        # Maximise la fenêtre
        driver.maximize_window()

        # Trouver l'élément "pe-cookies-simplified"
        pe_cookies_simplified = driver.find_element(By.CSS_SELECTOR, "body > pe-cookies-simplified")

        # Utiliser JavaScript pour accéder au shadow DOM et cliquer sur #pecookies-close
        script = 'arguments[0].shadowRoot.querySelector("#pecookies-close").click();'
        driver.execute_script(script, pe_cookies_simplified)
        print("Fenêtre pop-up fermée avec succès.")

        # Attendre 10 secondes avant de poursuivre
        time.sleep(10)

        # Identifiants de connexion
        username = "Sduchateau"
        password = "Steve95130!"

        # Remplir le nom d'utilisateur
        driver.find_element(By.ID, "LOGutilisateur").send_keys(username)

        # Remplir le mot de passe
        driver.find_element(By.ID, "LOGmotdepasse").send_keys(password)

        # Cliquer sur le bouton de connexion
        driver.find_element(By.ID, "BTconnecter").click()

        # Attendre 15 secondes
        time.sleep(15)

        # Cliquer sur l'application avec le lien contenant le texte "14521"
        application_link = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '14521')]")))
        application_link.click()

        # Attendre que la nouvelle fenêtre soit disponible
        time.sleep(15)

        # Basculer vers la nouvelle fenêtre
        driver.switch_to.window(driver.window_handles[1])

        # Basculer vers le cadre
        driver.switch_to.frame(0)

        # Cliquer sur les liens et boutons supplémentaires
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "ecranGlobal:menuHorizontal:j_id_id0pc4::_item1-input"))).click()
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "ecranGlobal:menuHorizontal:j_id_id0pc4::_item3-input"))).click()

        # Attendre un court instant avant de cliquer sur "Commande"
        time.sleep(2)

        # Cliquer sur "Commande"
        try:
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.LINK_TEXT, "Commande"))).click()
        except Exception as e:
            print("Erreur lors du clic sur 'Commande':", str(e))

        # Attendre que la nouvelle fenêtre soit disponible
        time.sleep(15)

        # Basculer vers la nouvelle fenêtre
        print("Nombre de fenêtres détectées:", len(driver.window_handles))

        # Basculer vers la nouvelle fenêtre
        if len(driver.window_handles) > 2:
            driver.switch_to.window(driver.window_handles[2])
        else:
            print("Pas assez de fenêtres détectées pour basculer.")

        # Cliquer sur "Critère de recherche"
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "ecranGlobal:ecran:ecranCorps:ongletCriteresDeRecherche::input"))).click()

        # Remplir le champ avec la lettre de commande actuelle
        lettre_commande = ws.cell(row=current_iteration + 1, column=1).value
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "ecranGlobal:ecran:ecranCorps:ongletCriteresDeRecherche:entryNumeroCommande"))).send_keys(
            lettre_commande)

        # Cliquer sur "Recherche"
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "ecranGlobal:ecran:ecranCorps:ongletCriteresDeRecherche:btnrechercher"))).click()

        # Attendre 10 secondes
        time.sleep(10)

        # Cliquer sur "Gérer les événements"
        try:
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.LINK_TEXT, "Gérer les événements"))).click()
        except Exception as e:
            print("Erreur lors du clic sur 'Gérer les événements':", str(e))

        # Attendre que la nouvelle fenêtre soit disponible
        time.sleep(10)

        # Basculer vers la nouvelle fenêtre
        if len(driver.window_handles) > 2:
            driver.switch_to.window(driver.window_handles[2])

        # Attendre 15 secondes
        time.sleep(15)

        # Utiliser BeautifulSoup pour extraire le prénom et le nom
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        prenom_nom_element = soup.find('p', {'id': 'rdv-prenom'})

        # Écrire les informations extraites dans la feuille Excel
        if prenom_nom_element:
            sheet.cell(row=current_iteration * 2 + 2, column=1, value=lettre_commande)
            sheet.cell(row=current_iteration * 2 + 2, column=2, value=prenom_nom_element.text.strip())

        # Extraire les données de la balise <span>
        span_elements = soup.find_all('span')

        # Écrire les données de toutes les balises <span> dans la même ligne
        col_index = 3
        for span_element in span_elements:
            sheet.cell(row=current_iteration * 2 + 2, column=col_index, value=span_element.text.strip())
            col_index += 1

        # Sauvegarder le fichier Excel après chaque itération
        workbook.save('extracted_data_combined.xlsx')

        # Sauvegarder l'itération réussie dans le fichier de sauvegarde
        with open(backup_file, 'w') as file:
            file.write(str(current_iteration))

    except Exception as ex:
        print(f"Une erreur s'est produite lors de l'itération {current_iteration}: {str(ex)}")

        # Enregistrez l'erreur dans le fichier Excel
        sheet.cell(row=current_iteration * 2 + 2, column=3, value=f"Erreur: {str(ex)}")

    finally:
        # Fermer le navigateur à la fin de chaque itération
        driver.quit()

    print(f"*** Fin de l'itération {current_iteration} ***\n")

    # Incrémenter le compteur d'itérations réussies
    last_successful_iteration = current_iteration

    # Attendre 1 minute 30 avant la prochaine itération
    time.sleep(90)

# Utiliser ThreadPoolExecutor pour exécuter les itérations en parallèle
with ThreadPoolExecutor(max_workers=7) as executor:  # Vous pouvez ajuster le nombre selon vos besoins
    futures = {executor.submit(process_iteration, i): i for i in range(start_iteration, max_iterations + 1)}

    for future in futures:
        try:
            future.result()
        except Exception as ex:
            print(f"Une erreur s'est produite: {str(ex)}")

# Imprimer un message une fois que toutes les itérations sont terminées
print("Toutes les itérations sont terminées. Consultez le fichier 'extracted_data_combined.xlsx' pour les résultats")
